"""
benchmark/cftr2_history.py — dating CFTR2's calls from its own release series
=============================================================================

Plumbing for the "when did CFTR2 first call this CF-causing?" section of
``benchmark/01_cftr2.ipynb``. The notebook keeps the *argument* — why a date is needed,
what the censoring floor does to it, which hold-outs it makes possible — and this module
holds the parts that would otherwise bury it: twelve workbooks with three different
column layouts, a nomenclature migration, a class vocabulary that changed twice, and the
guards that stop any of those from silently returning a plausible wrong answer.

Nothing here runs on import. CFTR2 keeps every past release at a stable public path, so
:func:`download_releases` can fetch the whole series; :func:`build_history` reads whatever
workbooks it is handed and raises if the series has a gap.

Why a date is measurable at all
-------------------------------
CFTR2's workbook has no per-variant date field — only a ``Date:`` row in the header. But
every release carries **two** determination columns, current and previous, and each
release's previous-version label names the immediately preceding release. Chained, the
twelve releases give a per-variant trajectory from 2015-08-13 to 2026-01-30, and the first
release at which a variant reads ``CF-causing`` is a real, CFTR2-native date.

What it cannot do
-----------------
The series is **left-censored at 2015-08-13**, the label on the 2016-08-08 workbook's
previous-version column: that is the oldest workbook cftr2.org serves (checked — no
2015-or-earlier xlsx exists at the path below). Everything already CF-causing at that
point has no date, only a bound. Resolution is set by CFTR2's irregular release cadence,
6-18 months.

The three traps this module exists to catch
-------------------------------------------
Each returns a plausible answer rather than an error, which is what makes them dangerous.

1. **The 2023 nomenclature migration.** 2023-04-07 moved to strict HGVS: ``c.1029delC``
   became ``c.1029del``, ``c.1021_1022dupTC`` became ``c.1021_1022dup``, and complex
   alleles gained bracket syntax. Keying the walk on cDNA name loses 140 variants at that
   one step and dates long-known alleles to 2023 — which for a leakage analysis is worse
   than no date, because it makes 2023-era tools look clean on variants known for decades.
   Legacy name is stable across all twelve releases and unique within each, so the walk
   keys on that; :func:`normalise_cdna` exists to corroborate, not to join.

2. **Class-vocabulary drift.** ``Unknown significance`` (2016-2023) became
   ``No interpretation available`` (2026). Treated as distinct values they manufacture a
   reclassification event for every affected variant, so :data:`CLASS_NORMAL` maps both to
   ``no_interpretation`` — while ``determination_raw`` keeps the exact source string.

3. **Rename tombstones.** 2024-09-25 introduced ``(CF-causing under new name)`` for seven
   rows whose previous column reads ``CF-causing (now renamed <new>)``. These are retired
   *old-name* entries, not classes. Scored as a determination they produce phantom
   downgrades; ignored entirely they orphan the successor, which then looks new in 2024
   despite being CF-causing for years. :func:`build_history` parses the rename target,
   aliases old to new, and merges the two trajectories.

Why this is a module and not more notebook cells
------------------------------------------------
Three header layouts, a nomenclature migration and a vocabulary remap take about two
hundred lines that are entirely about spreadsheet archaeology and say nothing about
cystic fibrosis. Inline, they drown the four lines that matter — read the releases, walk
them forward, write the extract, stamp the versions.

Why this is not in ``toolkit.py``
---------------------------------
``toolkit.py`` is the shared *reader* layer: one thin ``load_<tool>()`` per dataset, no
build logic, deliberately. ``load_cftr2_history()`` there reads what this module writes.

Why the ``_history`` suffix
---------------------------
Not decoration. A module named ``cftr2.py`` beside notebooks that run from ``benchmark/``
would sit on ``sys.path`` ahead of site-packages and shadow anything of that name — the
same failure ``tools/spliceai_build.py`` and ``tools/pangolin_build.py`` are named around.

Data use
--------
CFTR2's terms permit downloading for your own non-commercial use but forbid republishing
any portion of the Content, so neither the workbooks nor the extracts this writes may be
committed. Fetching them from cftr2.org is what the terms allow; redistributing them is
not. Cite CFTR2 (cftr2.org) if you use it.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

# --------------------------------------------------------------------------------------
# Vocabulary
# --------------------------------------------------------------------------------------
# Every determination string CFTR2 has used across the twelve releases, base form. Anything
# outside this set raises rather than being bucketed into a default -- a future CFTR2
# release that invents a class must fail loudly, not be silently miscounted.
CLASS_NORMAL = {
    "CF-causing": "cf_causing",
    "Varying clinical consequence": "varying",
    "Non CF-causing": "non_cf_causing",
    # The same concept, renamed between the 2023 and 2026 releases. Mapping them together
    # is what stops a pure relabelling from reading as a reclassification event.
    "Unknown significance": "no_interpretation",
    "No interpretation available": "no_interpretation",
}

CF_CAUSING = "cf_causing"

# The tombstone marker in a *current* column: this row is an old name being retired.
TOMBSTONE = "(CF-causing under new name)"
# Decorations wrapped around a base class in a determination cell. Both carry a rename;
# only the first names the successor, which is why the alias map is built from it.
_RENAMED_TO = re.compile(r"^(?P<base>.+?)\s*\(now renamed (?P<target>.+?)\)\s*$")
_RENAMED_FROM = re.compile(r"^(?P<base>.+?)\s*-\s*previously(?: called)? .+$")

# The censoring floor: the label on the 2016-08-08 workbook's previous-version column.
# Read from that file rather than hardcoded -- see build_history.
CENSORED = "left_censored"
OBSERVED = "observed"
NEVER = "never_cf_causing"

# CFTR2's own 2018-08-31 workbook disagrees with itself: its header block says
# CF-causing 336 / Varying 35, its rows say 335 / 36 (same 400 total). This is an upstream
# inconsistency, not a parse error -- verified by reading both directly. It is allowed here
# by explicit release + magnitude so that the cross-check still fails on anything else.
KNOWN_HEADER_DISCREPANCIES = {
    date(2018, 8, 31): {"CF-causing": -1, "Varying clinical consequence": 1},
}

_JUNK_PREFIXES = ("*", "©", "Permitted use", "Please use", "This detailed")

# --------------------------------------------------------------------------------------
# The release series
# --------------------------------------------------------------------------------------
# cftr2.org keeps every past variant-list release at a stable public path -- the same one
# its "CFTR2 Variant List History" page links. Verified 2026-08-19: each of these fetches
# 200 OK and hashes byte-identical to the locally archived copy, with no session cookie
# and without accepting the site usage agreement (that gate is on the variant *search*).
#
# The 2019 entry's filename really does contain a space and "(1)" on the server; it is not
# a local download artefact. The list is deliberately explicit rather than scraped: the
# history page is server-rendered behind a redirect, and a silently short list would just
# look like a shorter series. build_history's chain check is what proves none is missing.
CFTR2_RELEASE_BASE = "https://cftr2.org/sites/default/files"
CFTR2_RELEASE_FILES = (
    "CFTR2_8August2016.xlsx",
    "CFTR2_17March2017.xlsx",
    "CFTR2_8December2017_2.xlsx",
    "CFTR2_31August2018_3.xlsx",
    "CFTR2_11March2019 (1).xlsx",
    "CFTR2_10January2020.xlsx",
    "CFTR2_31July2020.xlsx",
    "CFTR2_24September2021.xlsx",
    "CFTR2_29April2022.xlsx",
    "CFTR2_7April2023.xlsx",
    "CFTR2_25September2024.xlsx",
    "CFTR2_30January2026.xlsx",
)


def download_releases(data_dir: Path, files=CFTR2_RELEASE_FILES) -> dict:
    """Fetch any release workbooks not already in ``data_dir``. Returns what it did.

    Downloading is what CFTR2's terms permit ("solely for your own non-commercial use");
    redistributing is not, which is why these land in gitignored ``data/`` and never in
    the repo. Existing files are left alone rather than re-fetched, so a local copy is
    never silently replaced by a differing upstream one.
    """
    import urllib.parse
    import urllib.request

    data_dir.mkdir(parents=True, exist_ok=True)
    got, already, failed = [], [], []
    for name in files:
        dest = data_dir / name
        if dest.exists():
            already.append(name)
            continue
        url = f"{CFTR2_RELEASE_BASE}/{urllib.parse.quote(name)}"
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                body = r.read()
            # A redirect to the agreement page returns HTML, not a workbook. Without this
            # the build would fail later with an opaque openpyxl zip error.
            if not body.startswith(b"PK"):
                raise HistoryError(f"{url} did not return an xlsx ({len(body)} bytes)")
            dest.write_bytes(body)
            got.append(name)
        except Exception as exc:  # noqa: BLE001
            failed.append((name, str(exc)[:120]))
    return {"downloaded": got, "already_present": already, "failed": failed}


class HistoryError(RuntimeError):
    """A guard fired. Every one of these means a date would otherwise be wrong."""


# --------------------------------------------------------------------------------------
# Parsing one release
# --------------------------------------------------------------------------------------
def _split_determination(cell) -> tuple[str | None, str | None, bool]:
    """Return ``(base_class, rename_target, is_tombstone)`` for one determination cell.

    CFTR2 decorates determinations rather than adding columns, so ``CF-causing (now
    renamed 3539del16)`` has to be split back into a class and a rename before either can
    be used. An empty cell means the variant had no previous determination, i.e. it was
    new in that release -- that is signal, not missing data, so it returns ``None``.
    """
    if cell is None:
        return None, None, False
    s = str(cell).strip()
    if not s or s in {"None", "-", "n/a"}:
        return None, None, False
    if s == TOMBSTONE:
        return None, None, True
    if (m := _RENAMED_TO.match(s)):
        return m.group("base").strip(), m.group("target").strip(), False
    if (m := _RENAMED_FROM.match(s)):
        return m.group("base").strip(), None, False
    return s, None, False


def _is_junk(legacy: str) -> bool:
    """Footnote and copyright lines sit in the same column as variant names.

    Four of them trail the 2024 workbook (``*Represents the allele frequency...``,
    ``©Copyright 2011 US CF Foundation...``). Parsed as variants they become four
    undatable rows with no determination.
    """
    return (not legacy) or legacy.startswith(_JUNK_PREFIXES) or len(legacy) > 60


def parse_release(path: Path) -> dict:
    """Read one CFTR2 workbook into ``{release_date, rows, header_counts, previous_label}``.

    Columns are located by **header text**, never by position: the 2016 release says
    "Mutation" where later ones say "Variant", and 2024 moved legacy name from column C to
    column A. Positional indexing works on exactly one layout and silently reads the wrong
    column on the other two.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # The header block is the preamble above the first row with several populated cells.
    try:
        hi = next(i for i, r in enumerate(rows) if sum(x is not None for x in r) >= 4)
    except StopIteration:
        raise HistoryError(f"{path.name}: no column-header row found")
    preamble, header = rows[:hi], [
        str(x).replace("\n", " ").strip() if x is not None else "" for x in rows[hi]
    ]

    release_date = None
    header_counts: dict[str, int] = {}
    for r in preamble:
        cell = str(r[0]).strip() if r[0] is not None else ""
        if cell.startswith("Date:"):
            release_date = datetime.strptime(cell[len("Date:"):].strip(), "%d %B %Y").date()
        # " CF-causing: 1,245" -- the workbook's own tally, used as a parse cross-check.
        if (m := re.match(r"^(.+?)\s*:\s*([\d,]+)$", cell)) and m.group(1).strip() in CLASS_NORMAL:
            header_counts[m.group(1).strip()] = int(m.group(2).replace(",", ""))
    if release_date is None:
        raise HistoryError(f"{path.name}: no 'Date:' row in the header block")

    def find(pattern: str) -> int:
        for j, h in enumerate(header):
            if re.search(pattern, h, re.I):
                return j
        raise HistoryError(f"{path.name}: no column matching {pattern!r} in {header}")

    i_legacy = find(r"(Variant|Mutation) legacy name")
    i_cdna = find(r"(Variant|Mutation) cDNA name")
    i_cur = find(r"final determination.*\(current")
    i_prev = find(r"final determination.*\(previous")

    out = []
    for r in rows[hi + 1:]:
        legacy = str(r[i_legacy]).strip() if r[i_legacy] is not None else ""
        if _is_junk(legacy):
            continue
        cur, renamed_to, tombstone = _split_determination(r[i_cur])
        prev, prev_target, _ = _split_determination(r[i_prev])
        out.append({
            "legacy_name": legacy,
            "cdna_name": str(r[i_cdna]).strip() if r[i_cdna] is not None else "",
            "current_raw": cur,
            "previous_raw": prev,
            # The tombstone's *previous* cell is the one naming the successor.
            "renamed_to": renamed_to or prev_target,
            "is_tombstone": tombstone,
        })

    dup = [k for k, n in Counter(row["legacy_name"] for row in out).items() if n > 1]
    if dup:
        raise HistoryError(
            f"{path.name}: legacy name is the join key but is not unique: {dup[:5]}")

    return {
        "path": path,
        "release_date": release_date,
        "rows": out,
        "header_counts": header_counts,
        "previous_label": header[i_prev],
        "current_label": header[i_cur],
    }


def normalise_cdna(name: str) -> str:
    """Best-effort cDNA name normalisation, for corroboration only -- never the join key.

    Absorbs the 2023 migration's two mechanical changes: trailing bases dropped from
    ``del``/``dup`` (``c.1029delC`` -> ``c.1029del``) and bracketed complex alleles
    (``c.[a;b]`` -> ``c.ab``). It does not attempt the irregular cases, which is precisely
    why the walk keys on legacy name instead.
    """
    if not name:
        return ""
    s = re.sub(r"\s+", "", str(name))
    s = re.sub(r"(del|dup|ins)[ACGT]+(?=$|[|;\]])", r"\1", s)
    s = s.replace("[", "").replace("]", "").replace(";", "")
    return s


# --------------------------------------------------------------------------------------
# Guards
# --------------------------------------------------------------------------------------
def check_chain(releases: list[dict]) -> list[str]:
    """Every release's previous-version label must name the preceding release's own Date:.

    A missing release does not announce itself — the walk simply skips forward and every
    variant first called CF-causing in the gap inherits the wrong, later date. This is the
    guard that makes "the archive is complete" a checked fact rather than an assumption.
    """
    notes = []
    for prev, cur in zip(releases, releases[1:]):
        label = cur["previous_label"]
        expect = prev["release_date"]
        # "Variant final determination 25 September 2024 (previous version)"
        m = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", label)
        if not m:
            raise HistoryError(f"{cur['path'].name}: cannot read a date from {label!r}")
        got = datetime.strptime(m.group(1), "%d %B %Y").date()
        if got != expect:
            raise HistoryError(
                f"release chain broken: {cur['path'].name} says its previous version is "
                f"{got} but the preceding file on disk is {expect}. A CFTR2 release is "
                f"missing from data/ -- every date after this point would be too late.")
        notes.append(f"{expect} -> {cur['release_date']}")
    return notes


def check_header_counts(rel: dict) -> None:
    """Per-release class tallies must match that workbook's own header block.

    The strongest end-to-end check available: it exercises the header-regex column
    mapping, the junk-row filter and the decoration stripping together, against a number
    CFTR2 computed itself.
    """
    if not rel["header_counts"]:
        return
    got = Counter()
    for row in rel["rows"]:
        if row["is_tombstone"] or row["current_raw"] is None:
            continue
        got[row["current_raw"]] += 1
    allowed = KNOWN_HEADER_DISCREPANCIES.get(rel["release_date"], {})
    for klass, expect in rel["header_counts"].items():
        delta = got[klass] - expect
        if delta != allowed.get(klass, 0):
            raise HistoryError(
                f"{rel['path'].name}: parsed {got[klass]} rows as {klass!r} but the "
                f"workbook header says {expect}. Column mapping or the junk-row filter "
                f"is wrong for this layout.")


def check_dropouts(steps: list[tuple[date, date, int, int]]) -> None:
    """No single release step may lose more than a handful of variants.

    This is the guard that actually catches a nomenclature migration, and the resurrection
    detector is not a substitute for it. When a key breaks, most affected variants do not
    vanish and return -- they vanish under the old key and reappear under a *new* one, so
    they read as one cohort dropped and another added, and every one of them is re-dated to
    the migration release. Only the handful that happen to revert trip a resurrection check.

    CFTR2 does retire variants, so dropouts are legitimate but rare. Measured across the
    twelve releases keyed on legacy name the worst step loses **4** of 1,167 (0.34%);
    keyed on normalised cDNA name the worst loses 39 of 400 (9.75%). The threshold below
    sits in the gap with room on both sides.
    """
    for prev_d, cur_d, present, dropped in steps:
        limit = max(5, present // 100)
        if dropped > limit:
            raise HistoryError(
                f"{dropped} of {present} variants present at {prev_d} are absent at "
                f"{cur_d} (limit {limit}). CFTR2 does not retire variants in bulk, so the "
                f"join key has almost certainly broken: those variants are still in the "
                f"list under a changed name and every one of them would be re-dated to "
                f"{cur_d}. Check the key column against that release's nomenclature.")


def check_vocabulary(releases: list[dict]) -> None:
    """No determination string may fall outside CLASS_NORMAL.

    Bucketing an unrecognised class into a default is how a future CFTR2 vocabulary change
    would quietly stop counting CF-causing calls.
    """
    unknown = Counter()
    for rel in releases:
        for row in rel["rows"]:
            for value in (row["current_raw"], row["previous_raw"]):
                if value is not None and value not in CLASS_NORMAL:
                    unknown[(rel["release_date"], value)] += 1
    if unknown:
        raise HistoryError(f"unrecognised determination values: {dict(unknown)}")


# --------------------------------------------------------------------------------------
# The walk
# --------------------------------------------------------------------------------------
def build_history(paths: list[Path], key_by: str = "legacy_name") -> dict:
    """Chain the releases into a per-variant determination trajectory.

    Returns ``{"long": DataFrame, "summary": DataFrame, "meta": dict}``. ``long`` is the
    raw uncollapsed signal, one row per variant per release; ``summary`` is the derived
    per-variant view the notebook and ``load_cftr2_history()`` use.

    ``key_by`` selects the join key and exists so the failure can be demonstrated rather
    than asserted. ``"legacy_name"`` is correct. ``"cdna_name"`` keys on the normalised
    cDNA name instead and is expected to trip the resurrection guard on the 2023
    nomenclature migration -- the notebook runs it deliberately to show the guard firing
    on the bug it is there to catch.
    """
    if key_by not in {"legacy_name", "cdna_name"}:
        raise HistoryError(f"key_by must be 'legacy_name' or 'cdna_name', got {key_by!r}")
    if not paths:
        raise HistoryError("no CFTR2 workbooks given")
    releases = sorted((parse_release(p) for p in paths), key=lambda r: r["release_date"])
    chain = check_chain(releases)
    check_vocabulary(releases)
    for rel in releases:
        check_header_counts(rel)

    # The oldest workbook's previous column is a real observation of an earlier state, but
    # an undated one -- CFTR2 names it only by the label on that column. It becomes a
    # pseudo-release marking the censoring floor.
    m = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", releases[0]["previous_label"])
    if not m:
        raise HistoryError(
            f"{releases[0]['path'].name}: cannot read the censoring floor from "
            f"{releases[0]['previous_label']!r}")
    floor = datetime.strptime(m.group(1), "%d %B %Y").date()

    # Rename aliases: old legacy name -> new legacy name, so the successor inherits the
    # predecessor's trajectory instead of looking new on the release that renamed it.
    alias: dict[str, str] = {}
    for rel in releases:
        for row in rel["rows"]:
            if row["renamed_to"]:
                alias[row["legacy_name"]] = row["renamed_to"]

    def resolve(name: str) -> str:
        seen, cur = set(), name
        while cur in alias and cur not in seen:
            seen.add(cur)
            cur = alias[cur]
        return cur

    def key_of(row: dict) -> str:
        if key_by == "cdna_name":
            return normalise_cdna(row["cdna_name"])
        return resolve(row["legacy_name"])

    # Per-release key sets, so a key break is caught before it becomes a date.
    key_sets = [(rel["release_date"],
                 {key_of(r) for r in rel["rows"] if not r["is_tombstone"]})
                for rel in releases]
    check_dropouts([(d1, d2, len(a), len(a - b))
                    for (d1, a), (d2, b) in zip(key_sets, key_sets[1:])])

    long_rows = []
    for i, rel in enumerate(releases):
        for row in rel["rows"]:
            if row["is_tombstone"]:
                continue  # the successor row carries this variant's history
            key = key_of(row)
            if i == 0 and row["previous_raw"] is not None:
                long_rows.append({
                    "variant_key": key, "release_date": floor,
                    "determination_raw": row["previous_raw"],
                    "determination": CLASS_NORMAL[row["previous_raw"]],
                    "cdna_name": row["cdna_name"], "censored_observation": True,
                })
            if row["current_raw"] is None:
                continue
            long_rows.append({
                "variant_key": key, "release_date": rel["release_date"],
                "determination_raw": row["current_raw"],
                "determination": CLASS_NORMAL[row["current_raw"]],
                "cdna_name": row["cdna_name"], "censored_observation": False,
            })

    long = (pd.DataFrame(long_rows)
            .sort_values(["variant_key", "release_date"])
            .reset_index(drop=True))

    latest = releases[-1]["release_date"]
    current_names = {key_of(r): r["cdna_name"]
                     for r in releases[-1]["rows"] if not r["is_tombstone"]}
    observed_dates = [r["release_date"] for r in releases]

    summary, resurrected = [], []
    for key, g in long.groupby("variant_key", sort=True):
        seen = list(dict.fromkeys(g["release_date"]))
        # A variant that vanishes and comes back was never really removed -- the key broke.
        # Contiguity is checked against the release calendar, not the row order.
        window = [d for d in ([floor] + observed_dates) if seen[0] <= d <= seen[-1]]
        if len(window) != len(seen):
            resurrected.append(key)
        cf = g.loc[g["determination"] == CF_CAUSING, "release_date"]
        first_cf = cf.min() if len(cf) else None
        if first_cf is None:
            basis = NEVER
        elif first_cf == floor:
            basis = CENSORED
        else:
            basis = OBSERVED
        # "Was CF-causing at some release and not CF-causing at a LATER one." Deliberately
        # not "is not CF-causing now": c.297-3C->T went CF-causing -> no interpretation ->
        # CF-causing, and an end-state test scores that round trip as never having wavered.
        # The pair (this, current_determination) answers both questions; a current-state
        # column alone cannot recover the interruption.
        withdrawn = bool(len(cf)) and (g["release_date"] > cf.min()).any() and (
            g.loc[g["release_date"] > cf.min(), "determination"] != CF_CAUSING).any()
        summary.append({
            "cdna_name": current_names.get(key, ""),
            "variant_key": key,
            "cftr2_first_cf_causing": first_cf,
            "cftr2_date_basis": basis,
            "cftr2_first_seen": seen[0],
            "cftr2_class_changes": int((g["determination"] != g["determination"].shift()).sum() - 1),
            "cftr2_ever_withdrawn": withdrawn,
            "cftr2_current_determination": g.iloc[-1]["determination"],
            "cftr2_in_current_release": key in current_names,
        })

    if resurrected:
        raise HistoryError(
            f"{len(resurrected)} variants disappear from the series and reappear "
            f"(e.g. {resurrected[:5]}). That is a join-key break, not a real removal: "
            f"their first-CF-causing dates would be too late. CFTR2 has probably changed "
            f"nomenclature again -- check the legacy-name column.")

    summary = pd.DataFrame(summary)
    missing = summary["cftr2_date_basis"].isna().sum()
    if missing:
        raise HistoryError(f"{missing} variants have no date basis")

    return {
        "long": long,
        "summary": summary,
        "meta": {
            "releases": [str(r["release_date"]) for r in releases],
            "release_count": len(releases),
            "chain": chain,
            "censoring_floor": str(floor),
            "latest_release": str(latest),
            "renames_applied": len(alias),
            "join_key": key_by,
            "source_workbooks": [r["path"].name for r in releases],
        },
    }


def write_extracts(hist: dict, data_dir: Path) -> dict:
    """Write the two extracts and the release stamp. All three are gitignored."""
    long_fp = data_dir / "cftr2_history_long.csv"
    summary_fp = data_dir / "cftr2_history.csv"
    release_fp = data_dir / "cftr2_history.release.json"

    hist["long"].to_csv(long_fp, index=False)
    summary = hist["summary"].copy()
    summary["cftr2_history_release"] = hist["meta"]["latest_release"]
    summary.to_csv(summary_fp, index=False)

    stamp = dict(hist["meta"])
    stamp["built_at_utc"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    stamp["row_count"] = len(summary)
    stamp["long_row_count"] = len(hist["long"])
    release_fp.write_text(json.dumps(stamp, indent=2))
    return {"long": long_fp, "summary": summary_fp, "release": release_fp}
